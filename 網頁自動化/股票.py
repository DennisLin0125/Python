from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet("大盤即時股價", 0)

wbTitle = ['類別', '股價', '漲跌', '漲跌幅', '開盤', '最高', '最低', '昨收']

ws.append(wbTitle)

category = {'台股': 2, '亞股': 3, '歐股': 4, '美股': 5}

Option = webdriver.ChromeOptions()
prefs = {
            "profile.default_content_setting_values":
            {
                "notifications": 2
            }
}
Option.add_experimental_option('prefs', prefs)

PATH = "./chromedriver"

url = f'https://tw.stock.yahoo.com/'

arr = []

driver = webdriver.Chrome(PATH, options=Option)
driver.get(url)

for data in category:
    arr.append(data)
    print(f'正在爬取 {data} 頁')

    get_ok = driver.find_elements(By.TAG_NAME, value="button")[category[data]]
    get_ok.click()

    pageSource = driver.page_source  # 取得網頁原始碼
    soup = BeautifulSoup(pageSource, "html.parser")  # 解析器接手

    全部資料 = soup.find_all('div', class_='Fx(a) Pt(22px) Pb(10px) Px(20px) Miw(200px)')[0]
    大盤 = 全部資料.contents[0].contents[0].text
    arr.append(大盤)

    if data == '台股':
        for i in 全部資料.next.contents[1].contents[1].contents:
            arr.append(i.text)
    else:
        for i in 全部資料.next.contents[1].contents[0].contents:
            arr.append(i.text)

    for all_data in 全部資料.contents[1].contents:
        arr.append(all_data.contents[1].text)

    股價 = arr[1]
    昨收 = arr[7]
    漲跌 = arr[2]
    漲跌幅 = arr[3]

    if float(股價) - float(昨收) > 0:
        arr[2] = f'△{漲跌}'
        arr[3] = f'△{漲跌幅.replace("(","").replace(")","")}'
    elif float(股價) - float(昨收) < 0:
        arr[2] = f'▽{漲跌}'
        arr[3] = f'▽{漲跌幅.replace("(","").replace(")","")}'

    ws.append(arr)
    arr = []

driver.quit()
#==========================================================================================
ws = wb.create_sheet("yahoo股價", 1)

wbTitle = ['類別', '代號', '股票名稱', '股價', '漲跌', '漲跌幅', '開盤', '昨收', '最高', '最低', '成交量(張)', '時間', '連結']

arrTable = [3, 6, 7, 8, 9, 10]

ws.append(wbTitle)

category = {
            '水泥': 1, '食品': 2, '塑膠': 3, '紡織': 4, '電機': 6, '電器電纜': 7, '化學': 37, '生技': 38,
            '玻璃': 9, '造紙': 10, '鋼鐵': 11, '橡膠': 12, '汽車': 13, '半導體': 40, '電腦週邊': 41, '光電': 42,
            '通訊網路': 43, '電子零組件': 44, '電子通路': 45, '資訊服務': 46, '其他電子': 47, '營建': 19, '航運': 20, '觀光': 21,
            '金融業': 22, '貿易百貨': 24, '油電燃氣': 39, '存託憑證': 25, 'ETF': 26, '受益證券': 29, 'ETN': 48, '創新板': 49,
            '其他': 30, '市認購': 31, '市認售': 32, '指數類': 33, '市牛證': 51, '市熊證': 52
}

arr = []
sumL = 0
for data in category:
    print(f'正在爬取 {data} 頁')
    url = f'https://tw.stock.yahoo.com/class-quote?sectorId={category[data]}&exchange=TAI'
    driver = webdriver.Chrome(PATH, options=Option)
    driver.get(url)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "p"))
    )

    temp = driver.find_element(By.TAG_NAME, value="p").text.replace('共 ','').replace(' 筆結果','')

    k = 0

    if int(temp) > 30:
        for i in range(10):  # 進行十次
            driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')  # 重複往下捲動
            time.sleep(1)

    pageSource = driver.page_source  # 取得網頁原始碼
    soup = BeautifulSoup(pageSource, "html.parser")  # 解析器接手

    for all_data in soup.find_all('li', class_='List(n)'):
        arr.append(data)
        連結 = all_data.find('a')['href']
        代號 = 連結.replace('https://tw.stock.yahoo.com/quote/','')
        arr.append(代號)
        for a in all_data.contents[0].contents:
            arr.append(a.text)
        k += 1
        arr.append(連結)
        股價 = arr[3].replace(',', '')
        昨收 = arr[7].replace(',', '')
        漲跌 = arr[4]
        漲跌幅 = arr[5]
        if 股價 != '-':
            if float(股價) - float(昨收) > 0:
                arr[4] = f'△{漲跌}'
                arr[5] = f'△{漲跌幅}'
            elif float(股價) - float(昨收) < 0:
                arr[4] = f'▽{漲跌}'
                arr[5] = f'▽{漲跌幅}'

        arr[2] = arr[2].replace(代號, '')

        for i in range(11):
            if i in arrTable and 股價 != '-':
                arr[i] = arr[i].replace(',', '')
                if i != 10:
                    arr[i] = float(arr[i])
                else:
                    arr[i] = int(arr[i])

        ws.append(arr)
        arr = []
    sumL += k
    print(f'共 {k} 筆')
    print('===============================')
    driver.quit()

wb.save('yahoo股價.xlsx')
print(f'全部 {sumL} 筆完成')
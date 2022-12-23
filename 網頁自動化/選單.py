from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet("大盤即時股價", 0)

wbTitle = ['類別', '股價', '漲跌', '漲跌幅', '開盤', '最高', '最低', '昨收']

ws.append(wbTitle)

category = {'台股': 2, '亞股': 3, '歐股': 4, '美股': 5}

Option = webdriver.ChromeOptions()
prefs = {
            "profile.default_content_setting_values":
            {
                "notifications": 2
            }
}
Option.add_experimental_option('prefs', prefs)

PATH = "./chromedriver"

url = f'https://tw.stock.yahoo.com/'

arr = []

driver = webdriver.Chrome(PATH, options=Option)
driver.get(url)

for data in category:
    arr.append(data)
    print(f'正在爬取 {data} 頁')

    get_ok = driver.find_elements(By.TAG_NAME, value="button")[category[data]]
    get_ok.click()

    pageSource = driver.page_source
    soup = BeautifulSoup(pageSource, "html.parser")

    全部資料 = soup.find_all('div', class_='Fx(a) Pt(22px) Pb(10px) Px(20px) Miw(200px)')[0]
    大盤 = 全部資料.contents[0].contents[0].text
    arr.append(大盤)

    if data == '台股':
        for i in 全部資料.next.contents[1].contents[1].contents:
            arr.append(i.text)
    else:
        for i in 全部資料.next.contents[1].contents[0].contents:
            arr.append(i.text)

    for all_data in 全部資料.contents[1].contents:
        arr.append(all_data.contents[1].text)

    股價 = arr[1]
    昨收 = arr[7]
    漲跌 = arr[2]
    漲跌幅 = arr[3]

    if float(股價) - float(昨收) > 0:
        arr[2] = f'△{漲跌}'
        arr[3] = f'△{漲跌幅.replace("(","").replace(")","")}'
    elif float(股價) - float(昨收) < 0:
        arr[2] = f'▽{漲跌}'
        arr[3] = f'▽{漲跌幅.replace("(","").replace(")","")}'

    ws.append(arr)
    arr = []
print('===============================')
driver.quit()
#==================================================================================
ws = wb.create_sheet("yahoo股價", 1)

wbTitle = ['類別', '代號', '股票名稱', '股價', '漲跌', '漲跌幅', '開盤', '昨收', '最高', '最低', '成交量(張)', '時間', '連結']

arrTable = [3, 6, 7, 8, 9, 10]

ws.append(wbTitle)

category = {
            '水泥': '食品',
            '食品': '塑膠',
            '塑膠': '紡織',
            '紡織': '電機',
            '電機': '電器電纜',
            '電器電纜': '化學',
            '化學': '生技',
            '生技': '玻璃',
            '玻璃': '造紙',
            '造紙': '鋼鐵',
            '鋼鐵': '橡膠',
            '橡膠': '汽車',
            '汽車': '半導體',
            '半導體': '電腦週邊',
            '電腦週邊': '光電',
            '光電': '通訊網路',
            '通訊網路': '電子零組件',
            '電子零組件': '電子通路',
            '電子通路': '資訊服務',
            '資訊服務': '其他電子',
            '其他電子': '營建',
            '營建': '航運',
            '航運': '觀光',
            '觀光': '金融業',
            '金融業': '貿易百貨',
            '貿易百貨': '油電燃氣',
            '油電燃氣': '存託憑證',
            '存託憑證': 'ETF',
            'ETF': '受益證券',
            '受益證券': 'ETN',
            'ETN': '創新板',
            '創新板': '其他',
            '其他': '指數類',
            '指數類': '水泥',
}

url = f'https://tw.stock.yahoo.com/class-quote?sectorId=1&exchange=TAI'
driver = webdriver.Chrome(PATH, options=Option)
driver.get(url)

arr = []
sumL = flag = k = 0

for data in category:

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.LINK_TEXT, f"上市類股 / {data}"))
    )
    get_ok = driver.find_element(By.LINK_TEXT, f"上市類股 / {data}")
    time.sleep(1)
    get_ok.click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.LINK_TEXT, category[data]))
    )

    ok = driver.find_element(By.LINK_TEXT, category[data])
    time.sleep(1)
    ok.click()

    time.sleep(3)

    temp = driver.find_element(By.TAG_NAME, value="p").text.replace('共 ', '').replace(' 筆結果', '')

    if int(temp) > 30:
        for j in range(10):
            flag = 1
            driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')  # 重複往下捲動
            time.sleep(1)
    if flag == 1:
        driver.execute_script("window.scrollTo(0,0)")
        time.sleep(1)
        flag = 0

    pageSource = driver.page_source
    soup = BeautifulSoup(pageSource, "html.parser")

    print(f'正在爬取 {category[data]} 頁')

    for all_data in soup.find_all('li', class_='List(n)'):
        arr.append(category[data])
        連結 = all_data.find('a')['href']
        代號 = 連結.replace('https://tw.stock.yahoo.com/quote/', '')
        arr.append(代號)
        for a in all_data.contents[0].contents:
            arr.append(a.text)
        k += 1
        arr.append(連結)
        股價 = arr[3].replace(',', '')
        昨收 = arr[7].replace(',', '')
        漲跌 = arr[4]
        漲跌幅 = arr[5]
        if 股價 != '-':
            if float(股價) - float(昨收) > 0:
                arr[4] = f'△{漲跌}'
                arr[5] = f'△{漲跌幅}'
            elif float(股價) - float(昨收) < 0:
                arr[4] = f'▽{漲跌}'
                arr[5] = f'▽{漲跌幅}'

        arr[2] = arr[2].replace(代號, '')

        for i in range(11):
            if i in arrTable and 股價 != '-':
                arr[i] = arr[i].replace(',', '')
                if i != 10:
                    arr[i] = float(arr[i])
                else:
                    arr[i] = int(arr[i])

        ws.append(arr)
        arr = []

    sumL += k
    print(f'共 {k} 筆')
    print('===============================')
    k = 0

wb.save('yahoo股價.xlsx')
driver.quit()
print(f'全部 {sumL} 筆完成')




